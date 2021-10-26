from io import BytesIO

import openpyxl
from django.utils.translation import ugettext_lazy as _


class ParserExcel:
    """
    Class parser excel file common
    Use:
        parserExcel = ParserExcel(<testfile>, <row_header = 0>, <row_body = 1>)
        parserExcel.parser()
        # get data
        data = parserExcel.excel_content
    """
    # Initializing index
    # define return
    RET_OK = 0
    RET_ERR = -1
    EXC_START_ROW_HEADER = 1
    EXC_START_ROW_BODY = 2
    TOTAL_ROWS = 0
    TOTAL_COLUMNS = 0
    WRN_HELP_COL_EXCEL = _("Total columns in file fail !")

    # Initializing variable protect
    _file_name = ''
    _excel_content = []

    def __init__(self, excel_file, row_start_header=1, row_start_body=2, number_col=- 1):
        """
        :param excel_file: file path
        :param row_start_header: row number read header
        :param row_start_body: row number read data
        :param number_col: total row read
        """
        self._file_name = excel_file
        self._number_col = number_col
        self.EXC_START_ROW_HEADER = row_start_header
        self.EXC_START_ROW_BODY = row_start_body

    def parser(self):
        file_name = self._file_name
        if not file_name:
            return 0

        try:
            wb_obj = openpyxl.load_workbook(BytesIO(file_name))
            sheet = wb_obj.active
            self.TOTAL_ROWS = sheet.max_row
            self.TOTAL_COLUMNS = sheet.max_column
            if self.TOTAL_COLUMNS != self._number_col + 1 and self._number_col != -1:
                return self.WRN_HELP_COL_EXCEL
        except Exception as ex:
            return str(ex)

        list_header = []
        list_body = []
        # parser header
        ret = self.parser_header(sheet, list_header, self.EXC_START_ROW_HEADER)
        if ret == self.RET_OK:
            # parser body
            self.parser_body(sheet, list_header, list_body, self.EXC_START_ROW_BODY)
        return list_header, list_body

    def parser_header(self, sheet, list_result, start_header_row):
        """
        :param sheet: obj sheet
        :param list_result: list header result
        :param start_header_row: row start reading
        :return:
        """
        for col_index in range(1, self.TOTAL_COLUMNS + 1):
            errcode = self.RET_OK
            header_data = sheet.cell(start_header_row, col_index).value
            if not header_data:
                errcode = self.RET_ERR
            tem_header = {
                "colIdx": col_index,
                "nameHeader": header_data,
                "errCode": errcode
            }
            list_result.append(tem_header)
        return self.RET_OK

    def parser_body(self, sheet, list_header, list_body, start_data_row):
        """
        :param sheet:
        :param list_header:
        :param list_body:
        :param start_data_row:
        :return:
        """
        for row_index in range(start_data_row, self.TOTAL_ROWS + 1):
            temp_row = []
            for index in range(len(list_header)):
                body_data = sheet.cell(row_index, list_header[index]["colIdx"]).value
                temp_row.append(body_data)
            list_body.append(temp_row)
        return self.RET_OK
